"""Clarksons-style Excel parser.

Handles the observed format:
  Row 1: [None, <series_id_1>, <series_id_2>, ...]
  Row 2: [None, <series_name_1>, <series_name_2>, ...]
  Row 3: ['Date', '$/day', '$/day', ...]
  Row 4+: [<excel_serial>, <value_1>, <value_2>, ...]

Data start row is detected by scanning column A for the first numeric cell in
the Excel serial-date range (15000–80000 ≈ 1941–2119).
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

import numpy as np
import openpyxl
import pandas as pd

from ..db import db_cursor

EXCEL_SERIAL_MIN = 15000
EXCEL_SERIAL_MAX = 80000

CLASS_PATTERNS: list[tuple[re.Pattern, str]] = [
    (re.compile(r"\bvlcc\b", re.I), "VLCC"),
    (re.compile(r"\bsuezmax\b", re.I), "SUEZMAX"),
    (re.compile(r"\baframax\b", re.I), "AFRAMAX"),
    (re.compile(r"\blr2\b", re.I), "LR2"),
    (re.compile(r"\blr1\b|panamax", re.I), "LR2"),
    (re.compile(r"\bmr\b|medium[- ]range", re.I), "MR"),
]


@dataclass
class ColumnSeries:
    column: int  # 1-based Excel column index
    series_id: Optional[str]
    series_name: Optional[str]
    unit: Optional[str]
    proposed_class: Optional[str]


def auto_match_class(name: Optional[str]) -> Optional[str]:
    if not name:
        return None
    for pat, code in CLASS_PATTERNS:
        if pat.search(name):
            return code
    return None


def excel_serial_to_date(serial: float) -> pd.Timestamp:
    """Convert Excel serial date to pandas Timestamp.

    Uses the 1900-date-system origin of 1899-12-30 (which handles the
    well-known Excel 1900 leap-year bug correctly for dates ≥ 1 Mar 1900).
    """
    return pd.to_datetime(serial, origin="1899-12-30", unit="D")


def preview_workbook(file_path: Path, sheet_name: str | int = 0) -> dict:
    """Peek at the raw first 10 rows and propose a class mapping.

    Used by the frontend before confirming ingest.
    """
    wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
    if isinstance(sheet_name, int):
        ws = wb.worksheets[sheet_name]
    else:
        ws = wb[sheet_name]

    raw_rows: list[list] = []
    for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
        raw_rows.append(list(row))

    # Need a slightly larger scan to find data start for accurate metadata rows
    serial_scan = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        a = row[0] if row else None
        if isinstance(a, (int, float)) and EXCEL_SERIAL_MIN <= a <= EXCEL_SERIAL_MAX:
            serial_scan.append(i)

    data_start_row = serial_scan[0] if serial_scan else None
    wb.close()

    columns: list[ColumnSeries] = []
    if data_start_row and data_start_row >= 2:
        # Row 1 → IDs; row (data_start_row-1) → units; row (data_start_row-2) → names
        units_row = raw_rows[data_start_row - 2] if data_start_row - 2 < len(raw_rows) else []
        names_row = raw_rows[data_start_row - 3] if data_start_row - 3 >= 0 and data_start_row - 3 < len(raw_rows) else []
        ids_row = raw_rows[0] if raw_rows else []
        ncols = max(len(ids_row), len(names_row), len(units_row))
        for c in range(1, ncols):  # skip col A (dates)
            sid = ids_row[c] if c < len(ids_row) else None
            sname = names_row[c] if c < len(names_row) else None
            unit = units_row[c] if c < len(units_row) else None
            columns.append(
                ColumnSeries(
                    column=c + 1,  # 1-based for display
                    series_id=str(sid) if sid is not None else None,
                    series_name=str(sname) if sname is not None else None,
                    unit=str(unit) if unit is not None else None,
                    proposed_class=auto_match_class(str(sname) if sname else None),
                )
            )

    return {
        "raw_rows": [
            [str(v) if v is not None else "" for v in r] for r in raw_rows[:10]
        ],
        "data_start_row": data_start_row,
        "columns": [c.__dict__ for c in columns],
    }


def ingest_clarksons_excel(
    file_path: Path,
    class_mapping: Optional[dict[int, str]] = None,
    sheet_name: str | int = 0,
) -> dict:
    """Ingest a Clarksons-style Excel file into tce_history.

    class_mapping: optional dict of {column_index_1based: vessel_class_code}
                   overrides or supplies when auto-detect fails.
    """
    wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
    ws = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]

    warnings: list[str] = []

    # Load everything we need into memory (files are small: ~1k rows)
    all_rows: list[list] = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    # Find data start row
    data_start = None
    for i, row in enumerate(all_rows, start=1):
        if not row:
            continue
        a = row[0]
        if isinstance(a, (int, float)) and EXCEL_SERIAL_MIN <= a <= EXCEL_SERIAL_MAX:
            data_start = i
            break

    if data_start is None:
        raise ValueError("Could not locate data start row — no Excel serial date found in column A.")

    ids_row = all_rows[0] if all_rows else []
    units_row = all_rows[data_start - 2] if data_start - 2 >= 0 else []
    names_row = all_rows[data_start - 3] if data_start - 3 >= 0 else []

    ncols = max(len(ids_row), len(units_row), len(names_row))
    series_cols: list[ColumnSeries] = []
    for c in range(1, ncols):
        sid = ids_row[c] if c < len(ids_row) else None
        sname = names_row[c] if c < len(names_row) else None
        unit = units_row[c] if c < len(units_row) else None
        proposed = auto_match_class(str(sname) if sname else None)
        series_cols.append(
            ColumnSeries(
                column=c + 1,
                series_id=str(sid) if sid is not None else None,
                series_name=str(sname) if sname else None,
                unit=str(unit) if unit else None,
                proposed_class=proposed,
            )
        )

    # Final class resolution per column
    column_to_class: dict[int, str] = {}
    for col in series_cols:
        override = (class_mapping or {}).get(col.column)
        if override and override.upper() != "SKIP":
            column_to_class[col.column] = override.upper()
        elif col.proposed_class:
            column_to_class[col.column] = col.proposed_class
        else:
            warnings.append(
                f"Column {col.column} (series '{col.series_name}') has no class mapping — skipped."
            )

    # Parse dates + values
    data_rows = all_rows[data_start - 1 :]
    records: list[tuple[str, str, float, str, str]] = []
    dates: list[pd.Timestamp] = []
    for row in data_rows:
        if not row or row[0] is None:
            continue
        a = row[0]
        if not isinstance(a, (int, float)):
            continue
        if not (EXCEL_SERIAL_MIN <= a <= EXCEL_SERIAL_MAX):
            continue
        dt = excel_serial_to_date(a)
        dates.append(dt)
        for col in series_cols:
            cls = column_to_class.get(col.column)
            if not cls:
                continue
            if col.column - 1 >= len(row):
                continue
            val = row[col.column - 1]
            if val is None:
                continue
            try:
                v = float(val)
            except (TypeError, ValueError):
                continue
            if not np.isfinite(v):
                continue
            records.append(
                (
                    dt.strftime("%Y-%m-%d"),
                    cls,
                    v,
                    col.series_id or "",
                    str(file_path.name),
                )
            )

    # Detect frequency
    median_days = None
    if len(dates) >= 3:
        deltas = np.diff(sorted(set(dates)))
        median_days = float(pd.Series(deltas).dt.days.median())
        if median_days == 7:
            freq = "weekly"
        elif median_days == 1:
            freq = "daily"
        elif 28 <= median_days <= 31:
            freq = "monthly"
        else:
            freq = f"{median_days:.0f}-day"
    else:
        freq = "unknown"

    # Data-quality warnings — surfaced in the UI before any calibration uses this data.
    if freq != "weekly" and median_days is not None:
        warnings.append(
            f"Detected {freq} frequency (median Δ = {median_days:.0f} d). "
            f"The weekly stochastic models may mis-specify — confirm the input."
        )
    # Per-class: short history + interior gaps
    by_class_dates: dict[str, list[pd.Timestamp]] = {}
    for (week_ending, cls, _v, _sid, _src) in records:
        by_class_dates.setdefault(cls, []).append(pd.to_datetime(week_ending))
    for cls, ds in by_class_dates.items():
        ds_sorted = sorted(set(ds))
        if len(ds_sorted) < 52:
            warnings.append(
                f"{cls}: only {len(ds_sorted)} observations (< 52 weeks). "
                f"Calibration may be unstable."
            )
        if median_days and len(ds_sorted) >= 3:
            class_deltas = np.diff(ds_sorted)
            class_days = pd.Series(class_deltas).dt.days.values
            gap_threshold = max(2 * median_days, median_days + 1)
            big_gaps = int((class_days > gap_threshold).sum())
            if big_gaps > 0:
                worst = int(class_days.max())
                warnings.append(
                    f"{cls}: {big_gaps} date gap(s) > {gap_threshold:.0f} d "
                    f"(largest = {worst} d). Interpolated TCE history may be missing periods."
                )

    # Insert with upsert semantics
    inserted = 0
    with db_cursor() as cur:
        for week_ending, cls, val, sid, src in records:
            cur.execute(
                """
                INSERT INTO tce_history (week_ending, vessel_class, route_code, tce_usd_per_day, source_series_id, source_file)
                VALUES (?, ?, '', ?, ?, ?)
                ON CONFLICT(week_ending, vessel_class, route_code)
                DO UPDATE SET tce_usd_per_day = excluded.tce_usd_per_day,
                              source_series_id = excluded.source_series_id,
                              source_file = excluded.source_file
                """,
                (week_ending, cls, val, sid, src),
            )
            inserted += cur.rowcount

    date_range = None
    if dates:
        date_range = (min(dates).strftime("%Y-%m-%d"), max(dates).strftime("%Y-%m-%d"))

    return {
        "series_found": [
            {
                "column": col.column,
                "series_id": col.series_id,
                "series_name": col.series_name,
                "unit": col.unit,
                "proposed_class": col.proposed_class,
                "matched": col.column in column_to_class,
            }
            for col in series_cols
        ],
        "n_observations": len(records),
        "date_range": date_range,
        "detected_frequency": freq,
        "rows_inserted": inserted,
        "warnings": warnings,
    }
