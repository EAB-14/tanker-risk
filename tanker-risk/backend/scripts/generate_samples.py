"""Generate two sample files that mimic the real data, for repo demonstration.

Produces:
  backend/data/samples/sample_clarksons.xlsx — same header layout as Clarksons
  backend/data/samples/sample.csv           — generic (date, vessel_class, tce)
"""
from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "data" / "samples"
OUT.mkdir(parents=True, exist_ok=True)


def synth_tce(n_weeks: int, mean: float, vol: float, seed: int) -> np.ndarray:
    """Weekly TCE path: mean-reverting OU with occasional jumps."""
    rng = np.random.default_rng(seed)
    x = np.empty(n_weeks)
    x[0] = mean
    kappa = 0.05
    for t in range(1, n_weeks):
        jump = 0.0
        if rng.uniform() < 0.02:
            jump = rng.normal(0, vol * 3)
        x[t] = max(5000.0, x[t - 1] + kappa * (mean - x[t - 1]) + rng.normal(0, vol) + jump)
    return x


def main() -> None:
    n_weeks = 520  # 10 years of weekly data
    start = date(2014, 1, 3)  # first Friday in Jan 2014
    dates = [start + timedelta(weeks=i) for i in range(n_weeks)]

    vlcc = synth_tce(n_weeks, mean=40000, vol=6000, seed=11)
    suez = synth_tce(n_weeks, mean=28000, vol=4500, seed=22)

    # ---- Clarksons-style Excel ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Row 1: series IDs
    ws.cell(1, 2, 99991)
    ws.cell(1, 3, 99992)
    # Row 2: names
    ws.cell(2, 2, "Average VLCC Long Run Historical Earnings (sample)")
    ws.cell(2, 3, "Average Suezmax Long Run Historical Earnings (sample)")
    # Row 3: units
    ws.cell(3, 1, "Date")
    ws.cell(3, 2, "$/day")
    ws.cell(3, 3, "$/day")
    # Row 4+: Excel-serial dates in col A
    for i, d in enumerate(dates):
        serial = (d - date(1899, 12, 30)).days
        ws.cell(4 + i, 1, serial)
        ws.cell(4 + i, 2, round(float(vlcc[i]), 2))
        ws.cell(4 + i, 3, round(float(suez[i]), 2))
    xlsx_path = OUT / "sample_clarksons.xlsx"
    wb.save(xlsx_path)
    print(f"Wrote {xlsx_path}")

    # ---- Generic CSV ----
    rows = []
    for i, d in enumerate(dates):
        rows.append({"date": d.isoformat(), "vessel_class": "VLCC", "tce": round(float(vlcc[i]), 2)})
        rows.append({"date": d.isoformat(), "vessel_class": "SUEZMAX", "tce": round(float(suez[i]), 2)})
    df = pd.DataFrame(rows)
    csv_path = OUT / "sample.csv"
    df.to_csv(csv_path, index=False)
    print(f"Wrote {csv_path}")


if __name__ == "__main__":
    main()
